import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from datetime import datetime, timedelta, time

START_ROW = 6
START_COL = 6
LAST_COL = None
ROW_STEP = 2
LATE_THRESHOLD = time(8, 30)
EARLY_THRESHOLD = time(18, 0)

FILL_FIRST = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_SECOND = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_THIRD = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
FONT_DEFAULT = Font(name="Times New Roman", size=10)
FONT_BOLD = Font(name="Times New Roman", size=10, bold=True)


def parse_time_value(val):
    if val is None: return None
    if isinstance(val, datetime): return val.time()
    if isinstance(val, time): return val
    s = str(val).strip()
    if s == "": return None
    if ":" in s:
        for fmt in ("%H:%M", "%H:%M:%S"):
            try: return datetime.strptime(s, fmt).time()
            except: pass
    try:
        f = float(val)
        if 0 <= f < 1:
            dt = datetime(1899, 12, 30) + timedelta(days=f)
            return dt.time()
    except: pass
    return None


def detect_last_col(ws):
    stop_keywords = ["giờ công","gio cong","ngày công","ngay cong","tăng ca","tang ca"]
    for r in range(1, 8):
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(row=r, column=c).value or "").strip().lower()
            if any(kw in val for kw in stop_keywords):
                return c - 1 if c > 1 else c
    return ws.max_column


def detect_name_col(ws):
    max_data, best_col = 0, 1
    for c in range(1, ws.max_column + 1):
        filled = 0
        for r in range(START_ROW, START_ROW + 10):
            val = str(ws.cell(row=r, column=c).value or "").strip()
            if val != "": filled += 1
        if filled > max_data:
            max_data, best_col = filled, c
    return best_col


def apply_color(cell, idx):
    if idx == 1:
        cell.fill = FILL_FIRST
    elif idx == 2:
        cell.fill = FILL_SECOND
    elif idx >= 3:
        cell.fill = FILL_THIRD


def set_border_range(ws, start_row, start_col, end_row, end_col):
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = border


def process_file(input_file, output_file):
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    last_col = LAST_COL or detect_last_col(ws)
    name_col = detect_name_col(ws)
    max_row = ws.max_row

    base_col = ws.max_column + 1
    headers = ["Lần 1", "Lần 2", "Lần 3", "Tổng"]

    ws.merge_cells(start_row=3, end_row=3, start_column=base_col, end_column=base_col + 3)
    cell_title = ws.cell(row=3, column=base_col, value="Tổng tiền phạt")
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    cell_title.font = FONT_BOLD
    set_border_range(ws, 3, base_col, 3, base_col + 3)

    for i, h in enumerate(headers):
        start_c = base_col + i
        ws.merge_cells(start_row=4, end_row=5, start_column=start_c, end_column=start_c)
        cell_top = ws.cell(row=4, column=start_c, value=h)
        cell_top.alignment = Alignment(horizontal="center", vertical="center")
        cell_top.font = FONT_BOLD
        cell_top.border = thin_border
        ws.cell(row=5, column=start_c).border = thin_border
        set_border_range(ws, 4, start_c, 5, start_c)

    r = START_ROW
    while r <= max_row:
        in_row, out_row = r, r + 1
        if out_row > max_row: break
        name_val = str(ws.cell(row=in_row, column=name_col).value or "").strip()
        if name_val == "":
            r += ROW_STEP
            continue
        violation_index = 0
        for c in range(START_COL, last_col + 1):
            cell_in = ws.cell(row=in_row, column=c)
            cell_out = ws.cell(row=out_row, column=c)
            cell_in.font = FONT_DEFAULT
            cell_out.font = FONT_DEFAULT
            t_in = parse_time_value(cell_in.value)
            t_out = parse_time_value(cell_out.value)
            if cell_in.value is None or str(cell_in.value).strip() == "":
                violation_index += 1; apply_color(cell_in, violation_index)
            elif t_in and LATE_THRESHOLD < t_in <= time(12, 0):
                violation_index += 1; apply_color(cell_in, violation_index)
            if t_out is None and t_in is not None:
                violation_index += 1; apply_color(cell_out, violation_index)
            elif t_out and time(15, 0) <= t_out < EARLY_THRESHOLD:
                violation_index += 1; apply_color(cell_out, violation_index)
        fine1 = 0
        fine2 = 50_000 if violation_index >= 2 else 0
        fine3 = 50_000 + (violation_index - 2) * 100_000 if violation_index >= 3 else 0
        total = fine1 + fine2 + fine3
        values = [fine1, fine2, fine3, total]
        for i, val in enumerate(values):
            c = base_col + i
            ws.merge_cells(start_row=in_row, end_row=out_row, start_column=c, end_column=c)
            cell = ws.cell(row=in_row, column=c, value=val)
            cell.number_format = "#,##0"
            cell.font = FONT_DEFAULT
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            set_border_range(ws, in_row, c, out_row, c)
        r += ROW_STEP
    wb.save(output_file)
